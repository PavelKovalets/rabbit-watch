"""Descriptive analytics over the visit log → an Excel workbook (FR-6).

    python -m src.brain.analytics [-o report.xlsx]

Reads the event log alone (reproducible from disk) and writes an `.xlsx` with sheets for
the raw visits, per-day counts, time-of-day distribution (local timezone), and a summary
incl. dwell-time stats. Descriptive only — no before/after efficacy comparison.
"""
import argparse
import statistics
from datetime import datetime

from src.brain.events import EventLog
from src.common.logger import get_logger

logger = get_logger("brain.analytics")


def _event_time(record: dict) -> str:
    """Visit start timestamp, tolerating the Phase 1 `timestamp` field name."""
    return record.get("start") or record.get("timestamp")


def _local(iso: str, tz=None) -> datetime:
    """Parse an ISO-8601 timestamp and convert to `tz` (default: system local)."""
    return datetime.fromisoformat(iso).astimezone(tz)


def visits_per_day(records, tz=None) -> dict[str, int]:
    counts: dict[str, int] = {}
    for r in records:
        day = _local(_event_time(r), tz).date().isoformat()
        counts[day] = counts.get(day, 0) + 1
    return counts


def visits_by_hour(records, tz=None) -> dict[int, int]:
    counts = {h: 0 for h in range(24)}
    for r in records:
        counts[_local(_event_time(r), tz).hour] += 1
    return counts


def dwell_stats(records) -> dict:
    durations = [r["duration_s"] for r in records if r.get("duration_s") is not None]
    if not durations:
        return {"count": 0, "mean": 0.0, "median": 0.0, "max": 0.0}
    return {
        "count": len(durations),
        "mean": statistics.mean(durations),
        "median": statistics.median(durations),
        "max": max(durations),
    }


def build_workbook(records, tz=None):
    """Build (but don't save) the analytics workbook."""
    from openpyxl import Workbook

    wb = Workbook()

    visits = wb.active
    visits.title = "Visits"
    visits.append(["start", "end", "duration_s", "confidence", "scene", "snapshot"])
    for r in sorted(records, key=_event_time):
        visits.append([
            _event_time(r), r.get("end", ""), r.get("duration_s", ""),
            r.get("confidence", ""), r.get("scene", ""), r.get("snapshot", ""),
        ])

    per_day = wb.create_sheet("Per day")
    per_day.append(["date", "visits"])
    for day, count in sorted(visits_per_day(records, tz).items()):
        per_day.append([day, count])

    by_hour = wb.create_sheet("By hour")
    by_hour.append(["hour", "visits"])
    hourly = visits_by_hour(records, tz)
    for hour in range(24):
        by_hour.append([hour, hourly[hour]])

    summary = wb.create_sheet("Summary")
    stats = dwell_stats(records)
    days = len(visits_per_day(records, tz))
    summary.append(["metric", "value"])
    summary.append(["total visits", len(records)])
    summary.append(["days observed", days])
    summary.append(["avg visits/day", round(len(records) / days, 2) if days else 0])
    summary.append(["dwell mean (s)", round(stats["mean"], 1)])
    summary.append(["dwell median (s)", round(stats["median"], 1)])
    summary.append(["dwell max (s)", stats["max"]])
    return wb


def write_report(out_path, records=None, tz=None) -> int:
    """Write the workbook to `out_path`; return the number of visits included."""
    records = EventLog().read_all() if records is None else records
    build_workbook(records, tz).save(out_path)
    return len(records)


def main():
    ap = argparse.ArgumentParser(description="Build an Excel analytics report from the visit log.")
    ap.add_argument("-o", "--out", default="data/brain/report.xlsx", help="output .xlsx path")
    args = ap.parse_args()
    n = write_report(args.out)
    print(f"wrote {args.out} ({n} visits)")


if __name__ == "__main__":
    main()
